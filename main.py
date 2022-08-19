from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer, LTChar
from openpyxl import Workbook


def main():

    wb = Workbook()
    ws = wb.active
    path = "logFile.pdf"
    flag = False
    row = 1
    for page_layout in extract_pages(path):
        for element in page_layout:
            if isinstance(element, LTTextContainer):
                for text_line in element:
                    # for character in text_line:
                    #     if isinstance(character, LTChar):

                    if str(text_line.get_text())[0:5] == 'After':
                        flag = True

                    if str(text_line.get_text())[0:2] == '==':
                        flag = False

                    if flag:
                        module_num = str(text_line.get_text())[1:7]
                        module_name = str(text_line.get_text())[9:]
                        ws['A' + str(row)] = module_num
                        ws['B' + str(row)] = module_name
                        row = row + 1

    wb.save("Log_Result2.xlsx")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
